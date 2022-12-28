import PyPDF2
import re
import glob
from openpyxl import workbook, load_workbook


xl_path = r'C:\Users\willd\OneDrive\Desktop\quote_compile\Excell\TEMPLATE.xlsx'
finished_location = r'C:\Users\willd\OneDrive\Desktop\quote_compile\Excell\quote_compile.xlsx'
wb = load_workbook(xl_path)

#folders
mack_quote_folder = r'C:\Users\willd\OneDrive\Desktop\quote_compile\mack_quotes'
core_quote_folder = r'C:\Users\willd\OneDrive\Desktop\quote_compile\core_quotes'
core_budget = r'C:\Users\willd\OneDrive\Desktop\quote_compile\CORE_BUDGET\part_budget'
ej_budget = r'C:\Users\willd\OneDrive\Desktop\quote_compile\CORE_BUDGET\casting_budget'

#excell sheets
pipe_sheet = wb['PIPE']
fitting_sheet = wb['FITTINGS']
mack_sheet = wb['MACK']
ej_sheet = wb['EAST JORDAN']
st_sheet = wb['STORMTECH']
br_sheet = wb['BLDG RISER']

#assorted lists
pipe_lines = []
fitting_lines = []
all_text = []
ej_item = []
ej_item_price = []
st_item = []
st_item_price = []
tech_list = []
br_size = []
br_price = []



# Makes a string of all text in all files, then assigns to correct list
def extract_core_text(file, suplier):
    global all_text
    for file in glob.glob(file + r"\*"):
        with open(file, 'rb') as pdfFile:
            reader = PyPDF2.PdfFileReader(pdfFile)
            number_of_pages = len(reader.pages)
            count = 0
            for pages in range(number_of_pages):
                page = reader.getPage(count)
                text = page.extractText()
                if re.search('CHAMBER', text):
                    line2 = text.replace(' ', '').splitlines()
                    # tech_list = []
                    for pos_total in line2:
                        if 'DETENTIONSYSTEM' and 'TOTAL' in pos_total:
                            pos_total = pos_total.split('TOTAL')[1]
                            if pos_total[0] == 'E':
                                pos_total = pos_total.replace(pos_total[0], '')
                                pos_total = pos_total.replace(pos_total[0], '')
                            if pos_total.count('.') > 1:
                                pos_total = pos_total.split('.', 1)[0]
                            job = file.split(r'core_quotes')[1]
                            job = job.replace(job[0], '')
                            tl1 = [job, pos_total]
                            tech_list.append(tl1)
                lines1 = text.splitlines()
                count += 1
                for line in lines1:
                    if suplier == 'c':
                        if re.search(" EA ", line):
                            fitting_lines.append(line)
                        elif re.search(" FT ", line):
                            pipe_lines.append(line)
                        elif 'BLDGRISERTOTAL' in line.replace(' ', ''):
                            # print(line)
                            building_riser(line)
                    else:
                        all_text.append(line)
    # storm_tech_xl(tech_list)
    # print(br_size)
    # print(br_price)

# creates a sheet of lump sum prices for building riser systems
def building_riser(pos_br):
    pos_br = pos_br.replace(' ', '').replace(',', '').rstrip()
    if pos_br[0] == '(':
        quant = pos_br.split(')')[0].replace('(', '')
        size1 = pos_br.split(')')[1]
        size = size1.split('TOTAL')[0]
        br_pr = float(pos_br.split('TOTAL')[1]) / int(quant)
    else:
        size = pos_br.split('TOTAL')[0]
        br_pr = pos_br.split('TOTAL')[1]
    br_size.append(size)
    br_price.append(br_pr)



# creates a sheet of lump sum prices for stormtech systems for further breakdow of unit price in excell
def storm_tech_xl(tech_list):
    storm_tech2 = wb['STORMTECH2']
    row = 1
    for syst in tech_list:
        row += 1
        storm_tech2['F' + str(row)] = syst[0]
        storm_tech2['J' + str(row)] = syst[1]


def mack_quote_clean(line1):
    global struct, price
    struct = []
    price = []
    for line in line1:
        if line[0].isdigit():
            if '$' in line:
                st1 = line.split(r' ', 1)[1]
                # st2 = st1.split('$')
                if len(st1.replace(' ', '')) > 4:
                    if st1[0] != '$':
                        if st1.count('$') > 1:
                            st_item = st1.split('$', 1)[0]
                            price_total = st1.split('$', 1)[1]
                            st_price = price_total.replace(' ', '').split('.')[0]
                            struct.append(str(st_item))
                            price.append(st_price)
                        # else:
                        #     st1 = st1.rstrip()
                        #     if st1[-1] != '$':
                        #         st_item = st1.split('$')[0]
                        #         st_price = st1.split('$')[1]
                        #         struct.append(str(st_item))
                        #         price.append(st_price)
                        #     else:
                        #         print(st1)
                            # struct.append(str(st_item))
                            # price.append(st_price)
    # print(len(struct), len(price))
    # print(struct)
    # print(price)


# extracts desired infor(product, price) creates two lists, one of products one of prices
def item_price_extract(unit, line_data):
    global fitting_item_price, pipe_item_price, fitting_item, pipe_item, item_actual
    pipe_item = []
    pipe_item_price = []
    fitting_item = []
    fitting_item_price = []
    for price_full in line_data:
        price_and_total = price_full.split(unit)[1]
        item_and_serial = price_full.split(unit)[0]
        price = price_and_total.split(r' ')[0]
        total_price = price_and_total.split(r' ')[1].replace(',', '')
        if price[0].isdigit():
            if float(total_price) != 0:
                quantity = float(total_price) / float(price.replace(',', ''))
                item_and_quantity = item_and_serial.split(r' ', 1)[1]
                test_quantity_value = item_and_quantity.split(r' ', 1)[0]
                if test_quantity_value == quantity:
                    item_actual = item_and_quantity.split(r' ', 1)[1]
                elif test_quantity_value != quantity:
                    item_actual = item_and_quantity[(len(str(int(quantity)))):]
                    item_actual = item_actual.lstrip()
                if item_actual[0] == "0":
                    item_actual = item_actual[1:]
                #     item_actual = item_actual.lstrip()
                if unit == ' FT ':
                    if 'N12' and 'WALL' in item_actual.replace(" ", ""):
                        item_actual = item_actual.replace(" ", "").split('WALL')[0] + ' WALL'
                    pipe_item.append(item_actual)
                    pipe_item_price.append(price)
                elif unit == ' EA ':
                    core_ea_cleanup(item_actual, price)


#clean up fitting items for beter specialization
def core_ea_cleanup(pos_fitting, pos_fitting_price):
    stormtech = ['chamber', 'stormtech', 'mc-3500', 'mc-720', 'mc720', 'sc-740', 'sc740', 'sc310', 'sc-160', 'sc160', 'mc3500', 'mc4500']
    pos_no_space = pos_fitting.replace(' ', '')
    if pos_fitting[0] == "E" and pos_fitting[1] == 'J':
        ej_item.append(pos_fitting)
        ej_item_price.append(pos_fitting_price)
    elif any(substring in pos_no_space.lower() for substring in stormtech):
        st_item.append(pos_fitting)
        st_item_price.append(pos_fitting_price)
    else:
        fitting_item.append(pos_fitting)
        fitting_item_price.append(pos_fitting_price)


# creates a dictionary of unique products and all corresponding prices
def product_price_dict(list2, list3):
    global dict1
    dict1 = {}
    for item in set(list2):
        indices = [i for i, x in enumerate(list2) if x == item]
        listx = []
        for index in indices:
            listx.append(list3[index])
        dict1[item] = listx
    # print(dict1)


# looks for keywords in products to assign material
def assign_material(item):
    global part_material
    pvc = ['PVC', 'SDR25', 'SDR35', 'MOLDED', 'CLEANOUT']
    clay = ['CLAY', 'LOGAN']
    water = ['NIP', 'TUB', 'COPPER', 'MJ', 'DI', 'FLG']
    storm = ['HDPE', 'N12', 'HP']
    if any(substring in item for substring in pvc):
        part_material = "PVC"
    elif any(substring in item for substring in storm):
        part_material = "STM"
    elif any(substring in item for substring in water):
        part_material = "water"
    elif any(substring in item for substring in clay):
        part_material = "CLAY"
    else:
        part_material = "MISC."


#structure type (manhole, basin, headwall)
def structure_type(item):
    global type
    hw = ['headwall', 'endwall']
    basin = ['basin', 'inlet']
    ocs = ['outletstructure', 'outletbasin']
    mh = ['manhole', 'mh', 'ocs']
    item = item.replace(' ', '').lower()
    if any(substring in item for substring in hw):
        type = "HW"
    elif any(substring in item for substring in basin):
        type = "CB"
    elif any(substring in item for substring in ocs):
        type = "ocs"
    elif any(substring in item for substring in mh):
        type = "MH"
    else:
        type = "MISC."



# transfers the created dictionaries to an excel document
def xl_transfer(dict_price, sheet, supplier, budget, budget_price):
    count3 = 1
    unique = []
    for x, prices_list in dict_price.items():
        count3 += 1
        count_cols = 7
        if supplier == 'c':
            assign_material(x)
            mat = part_material
            if x.replace(' ', '') in budget:
                index = budget.index(x.replace(' ', ''))
                xx = budget_price[index]
                del budget[index]
                del budget_price[index]
                # unique.append(x.replace(' ', ''))
                sheet['B' + str(count3)] = xx
        elif supplier == 'm':
            structure_type(x)
            mat = type
        else:
            mat = 'MISC.'
        sheet['G' + str(count3)] = x
        sheet['F' + str(count3)] = mat
        # sheet['B' + str(count3)] = xx
        for s in prices_list:
            count_cols += 1
            s = str(s).replace(',', '')
            sheet.cell(row=count3, column=count_cols).value = float(s)
    if supplier == 'c':
        for c4 in range(0, len(budget)):
            if budget[c4] not in unique:
                unique.append(budget[c4])
                assign_material(budget[c4])
                count3 += 1
                sheet['G' + str(count3)] = budget[c4]
                sheet['B' + str(count3)] = budget_price[c4]
                sheet['F' + str(count3)] = part_material
    storm_tech_xl(tech_list)
    wb.save(finished_location)


def budget_pricing(budget_folder):
    global bud_pip, bud_pip_pric, bud_fit_pric, bud_fit, bud_ej, bud_ej_pric
    extract_core_text(budget_folder, 'c')
    item_price_extract(" FT ", pipe_lines)
    bud_pip = [i.replace(' ', '') for i in pipe_item]
    bud_pip_pric = pipe_item_price
    # print(fitting_lines)
    item_price_extract(" EA ", fitting_lines)
    # print(fitting_item)
    bud_fit = [i.replace(' ', '') for i in fitting_item]
    bud_fit_pric = fitting_item_price
    bud_ej = [i.replace(' ', '') for i in ej_item]
    bud_ej_pric = ej_item_price

# budget_pricing(ej_budget)
# print(bud_fit)

budget_pricing(core_budget)
extract_core_text(mack_quote_folder, 'm')
mack_quote_clean(all_text)
product_price_dict(struct, price)
xl_transfer(dict1, mack_sheet, 'm', bud_fit, bud_fit_pric)
extract_core_text(core_quote_folder, 'c')
item_price_extract(" FT ", pipe_lines)
product_price_dict(pipe_item, pipe_item_price)
xl_transfer(dict1, pipe_sheet, 'c', bud_pip, bud_pip_pric)
item_price_extract(" EA ", fitting_lines)
product_price_dict(fitting_item, fitting_item_price)
xl_transfer(dict1, fitting_sheet, 'c', bud_fit, bud_fit_pric)
product_price_dict(ej_item, ej_item_price)
budget_pricing(ej_budget)
xl_transfer(dict1, ej_sheet, 'c', bud_ej, bud_ej_pric)
# product_price_dict(st_item, st_item_price)
# xl_transfer(dict1, st_sheet, 'st', bud_ej, bud_fit_pric)
product_price_dict(br_size, br_price)
xl_transfer(dict1, br_sheet, 'br', 7, 8)
